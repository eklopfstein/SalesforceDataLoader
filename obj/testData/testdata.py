# Standard packages
import datetime
import getpass
import logging
import signal
import sys
import tkinter
import tkinter.filedialog

# Community packages
from openpyxl import Workbook, load_workbook
from simple_salesforce import Salesforce, format_soql


def main(filePath = sys.argv[1], username = sys.argv[2], password = sys.argv[3], token = sys.argv[4], createUsers = sys.argv[5]):

    # registers handler for signal interrupt (i.e. Ctrl+C)
    signal.signal(signal.SIGINT, interruptHandler)
    # Will write logs to a file called result.log in the current directory (overwriting that file if it already exists)
    logging.basicConfig(filename="result.log", filemode="w", level=logging.DEBUG,
                        format='%(asctime)s - %(levelname)s: %(message)s')
    # Hides the root component for the GUI so it doesn't appear when no GUI is being used
    root = tkinter.Tk()
    root.withdraw()
    wb = loadWorkbook(filePath)
    sf = loginToSalesforce(username, password, token)
    users = getUsers(sf, wb, createUsers)
    parentAccounts = createParentAccounts(sf, users, wb)
    childAccounts = createChildAccounts(sf, users, parentAccounts, wb)
    personAccounts = createPersonAccounts(sf, users, wb)
    accounts = {**parentAccounts, **childAccounts, **personAccounts}
    contacts = createContacts(sf, users, wb)
    producers = createProducers(sf, users, wb, accounts, contacts)
    createLeads(sf, users, wb, accounts)
    createOpportunities(sf, users, wb, accounts)
    createTasks(sf, users, wb, accounts, contacts)
    createCases(sf, producers, wb, accounts, contacts)
    operatingHours = createOperatingHours(sf)
    workType = createWorkType(sf, operatingHours)
    serviceTerritory = createServiceTerritory(sf, operatingHours)
    createServiceTerritoryWorkType(sf, serviceTerritory, workType)
    createWorkTypeGroup(sf)
    wb.close()
    logInfo('Finished')


def interruptHandler(sig, frame):
    print("\nExiting program")
    sys.exit(0)


def logInfo(info):
    """Prints message and writes same message to log file

    Parameters:
        info (string) -- the message to output and write to log file

    Returns:
        void
    """
    print(info)
    logging.info(info)


def logError(errorMessage, error):
    """Prints error with indication to check log file for more information, prints full error to log file, then exits

    Parameters:
        errorMessage (string) -- the message to output to the user
        error (Exception) -- the error to output and write to log file

    Returns:
        void
    """
    print(errorMessage + ', check log file for more information')
    logging.info(errorMessage)
    logging.error(error)
    sys.exit(0)


def loadWorkbook(filePath):
    """Loads the test data workbook

    Returns:
        wb (openpyxl.workbook.Workbook) -- The workbook containing the test data to create
    """
    fileTypes = (("Excel files", "*.xlsx"), ("All Files", "*.*"))
    try:
        logInfo("Prompting for Excel workbook")
        # Opens a file dialog box and then opens the Excel workbook selected
        wb = load_workbook(filePath)
        logInfo("Loading Excel workbook")
        return wb
    except Exception as ex:
        logError("Could not load Excel workbook", ex)


def loginToSalesforce(uname, pas, token):
    """Gets user credentials and logs into Salesforce

    Returns:
        sf (Salesforce) -- The Salesforce session object
    """
    try:
        logInfo("Logging into Salesforce")
        sf = Salesforce(username=uname, password=pas,
                            security_token=token, domain='test')  # domain='test' means we're logging into a sandbox
        logInfo("Logged in")
        return sf
    except Exception as ex:
        logError("Could not log in", ex)


def createRecordMap(records, sobject):
    """Creates a map/dict to get the record ids by the record name

    Parameters:
        records (dict of string : list of dict of string : string) -- the records to create the map from
        sobject (string) -- the Salesforce object the records are of

    Returns:
        recordMap (dict of string : string) -- A dictionary with the key being the name of the record and the value being the id
    """
    logInfo("Creating " + sobject + " map")
    recordMap = dict()
    for r in records.get('records'):
        recordMap[r.get('Name')] = r.get('Id')
    logInfo("Created " + sobject + " map")
    return recordMap


def queryCreatedRecords(sf, records, sobject):
    """Queries for the name and id of the records that are passed in

    Parameters:
        sf (Salesforce) -- the active Salesforce connection
        records (list of dict of string : string) -- a list containing one item which is a dictionary of information regarding the created records
        sobject (string) -- the Salesforce object these are records of

    Returns:
        createRecordMap (queriedRecords, sobject) (dict of string : string) -- A dictionary of the queried records with the name as the key and the id as the value
    """
    try:
        recordIds = []
        for r in records[0]:
            recordIds.append(r.get('id'))
        logInfo("Querying created " + sobject + "s")
        q = "SELECT Id, Name FROM " + sobject + " WHERE Id IN {ids}"
        q = format_soql(q, ids=recordIds)
        logInfo("Query: " + q)
        queriedRecords = sf.query(q)
        logInfo("Got created " + sobject + " records")
        return createRecordMap(queriedRecords, sobject)
    except Exception as ex:
        logError("Could not query " + sobject + "s", ex)


def getRecordTypes(sf, ws, col, sobject):
    """Gets the record types for a given object

    Parameters:
        sf (Salesforce) -- the active Salesforce connection
        ws (openpyxl.workbook.Worksheet) -- the worksheet to read the record type names from
        col (integer) -- the record type column in the worksheet
        sobject (string) -- the object whose record types need queried

    Returns:
        createRecordMap(recordTypes, sobject) (dict of string : string>) -- A dictionary of record types for the given object
    """
    recordTypeNames = set()
    try:
        logInfo("Getting " + sobject.lower() + " record type names from worksheet")
        for recordType in ws.iter_rows(min_row=2, min_col=col, max_col=col, values_only=True):
            recordTypeNames.add(recordType[0])
    except Exception as ex:
        logError("Could not read " + sobject.lower() + " record type names from worksheet", ex)

    try:
        logInfo("Querying " + sobject.lower() + " record types")
        recordTypes = sf.query(
            format_soql(
                "SELECT Id, Name FROM RecordType WHERE SobjectType = {obj} AND IsActive = TRUE AND Name in {names}",
                obj=sobject, names=list(recordTypeNames)))
        logInfo("Got " + sobject.lower() + " record types")
        return createRecordMap(recordTypes, sobject)
    except Exception as ex:
        logError("Could not query " + sobject.lower() + " record types", ex)


def getUsers(sf, wb, createUsers):
    """Checks if the user wants to create users or not. If not, queries existing users instead

    Parameters:
        sf (Salesforce) -- the active Salesforce connection
        wb (openpyxl.workbook.Workbook) -- The workbook containing the test data to create

    Returns:
        (dict of string : string) -- A dictionary of users where the name is the key and the Id is the value
    """
    createOrQuery = None
    if (createUsers.lower() == "true"):
        return createUsers(sf, wb)
    else:
        return queryUsers(sf, wb)


def queryUsers(sf, wb):
    ws = wb["Users"]
    userNames = []
    try:
        logInfo("Reading users from Excel")
        for name in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            if (name[0] != None and name[1] != None):
                userNames.append(name[0] + " " + name[1])
    except Exception as ex:
        logError("Could not read users", ex)

    try:
        logInfo("Querying users")
        users = sf.query(format_soql("SELECT Id, Name FROM User WHERE Name IN {names}", names=userNames))
        return createRecordMap(users, "User")
    except Exception as ex:
        logError("Could not query users", ex)


def createUsers(sf, wb):
    """Creates users in the target org

    Parameters:
        sf (Salesforce) -- the active Salesforce connection
        wb (openpyxl.workbook.Workbook) -- The workbook containing the test data to create

    Returns:
        user (dict of string : string) -- a dictionary of the created users with Name as the key and Id as the value
    """
    try:
        logInfo("Getting user to set username")
        instance = sf.sf_instance
        orgName = '.' + instance.replace('.my.salesforce.com', '').replace('westernsouthernfinancialgroup--', '')
    except Exception as ex:
        logError("Could not get user", ex)

    ws = wb["Users"]
    try:
        logInfo("Querying profiles")
        profileNames = set()
        for profile in ws.iter_rows(min_row=2, min_col=6, max_col=6, values_only=True):
            profileNames.add(profile[0])
        profiles = sf.query(format_soql(
            "SELECT Id, Name FROM Profile WHERE Name IN {pro}", pro=list(profileNames)))
        profileMap = createRecordMap(profiles, "Profile")
    except Exception as ex:
        logError("Could not query profiles", ex)

    try:
        logInfo("Querying roles")
        roleNames = set()
        for role in ws.iter_rows(min_row=2, min_col=7, max_col=7, values_only=True):
            roleNames.add(role[0])
        roles = sf.query(format_soql("SELECT Id, Name FROM UserRole WHERE Name IN {rol}", rol=list(roleNames)))
        roleMap = createRecordMap(roles, "Role")
    except Exception as ex:
        logError("Could not query roles", ex)

    try:
        logInfo("Reading users from Excel")
        insertUsers = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if (row[0] == None):
                continue
            insertUsers.append(
                {'FirstName': row[0],
                 'LastName': row[1],
                 'Username': str(row[2]) + orgName, 'Email': row[3],
                 'Title': row[4],
                 'ProfileId': profileMap.get(row[5]),
                 'UserRoleId': roleMap.get(row[6]),
                 # creates alias from first character of first name and the first seven characters of the last name. Slicing a string shorter than seven characters doesn't cause an out of bounds exception
                 'Alias': str(row[0][0]) + str(row[1][0:7]),
                 'IsActive': True, 'TimeZoneSidKey': 'America/New_York', 'LocaleSidKey': 'en_US',
                 'EmailEncodingKey': 'UTF-8', 'LanguageLocaleKey': 'en_US'})
        logInfo("Creating users")
        users = sf.bulk.User.insert(insertUsers, batch_size=100)
        logInfo("Created users")
        logging.info(users)
        return queryCreatedRecords(sf, users, "User")
    except Exception as ex:
        logError("Could not create users", ex)


def createParentAccounts(sf, users, wb):
    """Creates parent accounts in the target org

    Parameters:
        sf (Salesforce) -- the active Salesforce connection
        users (dict of string : string) -- the users to assign ownership of the parent accounts to
        wb (openpyxl.workbook.Workbook) -- the workbook containing the test data to create

    Returns:
        parentAccounts (dict of string : string) -- a dictionary of the created parent accounts with Name as the key and Id as the value
    """
    ws = wb["ParentAccounts"]  # Gets the ParentAccounts sheet
    recordTypeMap = getRecordTypes(sf, ws, 3, "Account")
    insertParentAccounts = []
    try:
        logInfo("Reading Parent Accounts from Excel")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if (row[0] == None):
                continue
            insertParentAccounts.append(
                {'Name': row[0],
                 'EEP_Legal_Name_Of_Business__c': u"" if row[1] is None else row[1], 'RecordTypeId': recordTypeMap.get(
                     row[2]),
                 'OwnerId': users.get(row[3]),
                 'BillingStreet': u"" if row[4] is None else row[4], 'BillingCity': u"" if row[5] is None else row
                 [5], 'BillingState': u"" if row[6] is None else row[6], 'BillingPostalCode': u""
                 if row[7] is None else row[7], 'BillingCountry': u"" if row[8] is None else row[8], 'Phone': u""
                 if row[9] is None else row[9], 'EEP_Other_Phone__c': u"" if row[10] is None else row[10], 'Fax': u""
                 if row[11] is None else row[11], 'EEP_Restricted_Access__c': u"" if row[12] is None else row[12],
                 'EEP_Producer_Account_Tax_Id__c': u"" if row[14] is None else row[14], 'Website': u""
                 if row[15] is None else row[15], 'NumberOfEmployees': u"" if row[16] is None else row[16],
                 'FinServ__ClientCategory__c': u"" if row[17] is None else row[17], 'FinServ__Status__c': u""
                 if row[18] is None else row[18], 'FinServ__PersonalInterests__c': u"" if row[19] is None else row
                 [19], 'FinServ__MarketingSegment__c': u"" if row[20] is None else row[20],
                 'FinServ__FinancialInterests__c': u"" if row[21] is None else row[21], 'FinServ__ServiceModel__c': u""
                 if row[22] is None else row[22], 'FinServ__ReviewFrequency__c': u"" if row[23] is None else row[23],
                 'FinServ__InvestmentExperience__c': u"" if row[24] is None else row[24],
                 'FinServ__InvestmentObjectives__c': u"" if row[25] is None else row[25]})
        logging.info(insertParentAccounts)
    except Exception as ex:
        logError("Could not read Parent Accounts", ex)
    try:
        logInfo("Creating Parent Accounts")
        parentAccounts = sf.bulk.Account.insert(insertParentAccounts, batch_size=100)
        logInfo("Created Parent Accounts")
        logging.info(parentAccounts)
        return queryCreatedRecords(sf, parentAccounts, "Account")
    except Exception as ex:
        logError("Could not create Parent Accounts", ex)


def createChildAccounts(sf, users, parentAccounts, wb):
    """Creates child accounts in the target org

    Parameters:
        sf (Salesforce) -- the active Salesforce connection
        users (dict of string : string) -- the users to assign ownership of the accounts to
        parentAccounts (dict of string: string) -- the accounts that will have child accounts in test data
        wb (openpyxl.workbook.Workbook) -- the workbook containing the test data to create

    Returns:
        childAccounts (dict of string : string) -- a dictionary of the created accounts with Name as the key and Id as the value
    """
    ws = wb["ChildAccounts"]  # Gets the ChildAccounts sheet
    recordTypeMap = getRecordTypes(sf, ws, 3, "Account")
    insertChildAccounts = []
    try:
        logInfo("Reading Child Accounts from Excel")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if (row[0] == None):
                continue
            insertChildAccounts.append(
                {'Name': row[0],
                 'EEP_Legal_Name_Of_Business__c': u"" if row[1] is None else row[1], 'RecordTypeId': recordTypeMap.get(
                     row[2]),
                 'OwnerId': users.get(row[3]),
                 'BillingStreet': u"" if row[4] is None else row[4], 'BillingCity': u"" if row[5] is None else row
                 [5], 'BillingState': u"" if row[6] is None else row[6], 'BillingPostalCode': u""
                 if row[7] is None else row[7], 'BillingCountry': u"" if row[8] is None else row[8], 'Phone': u""
                 if row[9] is None else row[9], 'EEP_Other_Phone__c': u"" if row[10] is None else row[10], 'Fax': u""
                 if row[11] is None else row[11], 'EEP_Restricted_Access__c': u"" if row[12] is None else row[12],
                 'ParentId': parentAccounts.get(row[13]),
                 'EEP_Producer_Account_Tax_Id__c': u"" if row[14] is None else row[14], 'Website': u""
                 if row[15] is None else row[15], 'NumberOfEmployees': u"" if row[16] is None else row[16],
                 'FinServ__ClientCategory__c': u"" if row[17] is None else row[17], 'FinServ__Status__c': u""
                 if row[18] is None else row[18], 'FinServ__PersonalInterests__c': u"" if row[19] is None else row
                 [19], 'FinServ__MarketingSegment__c': u"" if row[20] is None else row[20],
                 'FinServ__FinancialInterests__c': u"" if row[21] is None else row[21], 'FinServ__ServiceModel__c': u""
                 if row[22] is None else row[22], 'FinServ__ReviewFrequency__c': u"" if row[23] is None else row[23],
                 'FinServ__InvestmentExperience__c': u"" if row[24] is None else row[24],
                 'FinServ__InvestmentObjectives__c': u"" if row[25] is None else row[25]})
        logging.info(insertChildAccounts)
    except Exception as ex:
        logError("Could not read Child Accounts", ex)

    try:
        logInfo("Creating Child Accounts")
        childAccounts = sf.bulk.Account.insert(insertChildAccounts, batch_size=100)
        logInfo("Created Child Accounts")
        logging.info(childAccounts)
        return queryCreatedRecords(sf, childAccounts, "Account")
    except Exception as ex:
        logError("Could not create Child Accounts", ex)


def createPersonAccounts(sf, users, wb):
    """Creates person accounts in the target org

    Parameters:
        sf (Salesforce) -- the active Salesforce connection
        users (dict of string : string) -- the users to assign ownership of the accounts to
        wb (openpyxl.workbook.Workbook) -- the workbook containing the test data to create

    Returns:
        personAccounts (dict of string : string) -- a dictionary of the created person accounts with Name as the key and Id as the value
    """
    ws = wb["PersonAccounts"]  # Gets the PersonAccounts sheet
    recordTypeMap = getRecordTypes(sf, ws, 3, "Account")
    insertPersonAccounts = []
    try:
        logInfo("Reading Person Accounts from Excel")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if (row[0] == None):
                continue
            insertPersonAccounts.append(
                {'EEP_Legal_Name_Of_Business__c': u"" if row[1] is None else row[1], 'RecordTypeId': recordTypeMap.get(
                    row[2]),
                 'OwnerId': users.get(row[3]),
                 'BillingStreet': u"" if row[4] is None else row[4], 'BillingCity': u"" if row[5] is None else row
                 [5], 'BillingState': u"" if row[6] is None else row[6], 'BillingPostalCode': u""
                 if row[7] is None else row[7], 'BillingCountry': u"" if row[8] is None else row[8], 'Phone': u""
                 if row[9] is None else row[9], 'EEP_Other_Phone__c': u"" if row[10] is None else row[10], 'Fax': u""
                 if row[11] is None else row[11], 'EEP_Restricted_Access__c': u"" if row[12] is None else row[12],
                 'EEP_Producer_Account_Tax_Id__c': u"" if row[14] is None else row[14], 'Website': u""
                 if row[15] is None else row[15], 'NumberOfEmployees': u"" if row[16] is None else row[16],
                 'FinServ__ClientCategory__c': u"" if row[17] is None else row[17], 'FinServ__Status__c': u""
                 if row[18] is None else row[18], 'FinServ__PersonalInterests__c': u"" if row[19] is None else row
                 [19], 'FinServ__MarketingSegment__c': u"" if row[20] is None else row[20],
                 'FinServ__FinancialInterests__c': u"" if row[21] is None else row[21], 'FinServ__ServiceModel__c': u""
                 if row[22] is None else row[22], 'FinServ__ReviewFrequency__c': u"" if row[23] is None else row[23],
                 'FinServ__InvestmentExperience__c': u"" if row[24] is None else row[24],
                 'FinServ__InvestmentObjectives__c': u"" if row[25] is None else row[25], 'Salutation': u""
                 if row[26] is None else row[26], 'FirstName': u"" if row[27] is None else row[27], 'LastName': u""
                 if row[28] is None else row[28], 'MiddleName': u"" if row[29] is None else row[29], 'Suffix': u""
                 if row[30] is None else row[30], 'PersonEmail': u"" if row[31] is None else row[31], 'Industry': u""
                 if row[32] is None else row[32]})
        logging.info(insertPersonAccounts)
    except Exception as ex:
        logError("Could not read Person Accounts", ex)

    try:
        logInfo("Creating Person Accounts")
        personAccounts = sf.bulk.Account.insert(insertPersonAccounts, batch_size=100)
        logInfo("Created Person Accounts")
        logging.info(personAccounts)
        return queryCreatedRecords(sf, personAccounts, "Account")
    except Exception as ex:
        logError("Could not create Person Accounts", ex)


def createContacts(sf, users, wb):
    """Creates contacts in the target org

    Parameters:
        sf (Salesforce) -- the active Salesforce connection
        users (dict of string : string) -- the user to assign ownership of the contacts to
        wb (openpyxl.workbook.Workbook) -- the workbook containing the test data to create

    Returns:
        contacts (dict of string : string) -- a dictionary of the created contacts with Name as the key and Id as the value
    """
    ws = wb["Contacts"]  # Gets the Contacts sheet
    recordTypeMap = getRecordTypes(sf, ws, 3, "Contact")
    insertContacts = []
    try:
        logInfo("Reading Contacts")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if (row[0] == None):
                continue
            insertContacts.append({'FirstName': row[0], 'LastName': row[1], 'RecordTypeId': recordTypeMap.get(
                row[2]), 'OwnerId': users.get(row[3])})
            logging.info(insertContacts)
    except Exception as ex:
        logError("Could not read Contacts", ex)

    try:
        logInfo("Creating Contacts")
        contacts = sf.bulk.Contact.insert(insertContacts, batch_size=100)
        logInfo("Created Contacts")
        logging.info(contacts)
        return queryCreatedRecords(sf, contacts, "Contact")
    except Exception as ex:
        logError("Could not create Contacts", ex)


def createProducers(sf, users, wb, accounts, contacts):
    """Creates producers in the target org

    Parameters:
        sf (Salesforce) -- the active Salesforce connection
        users (dict of string : string) -- the user to assign ownership of the producers to
        wb (openpyxl.workbook.Workbook) -- the workbook containing the test data to create
        accounts (dict of string : string) -- the accounts to relate the producers to
        contacts (dict of string : string) -- the contacts to relate the producers to

    Returns:
        producers (dict of string : string) -- a dictionary of the created producers with Name as the key and Id as the value
    """
    ws = wb["Producers"]
    insertProducers = []
    try:
        logInfo("Reading Producers")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if (row[0] == None):
                continue
            insertProducers.append(
                {'Name': row[0],
                 'AccountId': accounts.get(row[1]),
                 'ContactId': contacts.get(row[2]),
                 # converts the date into a standardized datetime string then removes the time part due to the field only being a date field
                 'EEP_Producer_Contract_Date__c': str(row[3].isoformat()).replace('T00:00:00', ''),
                 'EEP_Producer_Id__c': row[4],
                 'OwnerId': users.get(row[5])})
            logging.info(insertProducers)
    except Exception as ex:
        logError("Could not read Producers", ex)

    try:
        logInfo("Creating Producers")
        producers = sf.bulk.Producer.insert(insertProducers, batch_size=100)
        logInfo("Created Producers")
        logging.info(producers)
        return queryCreatedRecords(sf, producers, "Producer")
    except Exception as ex:
        logError("Could not create Producers", ex)


def createLeads(sf, users, wb, accounts):
    """Creates Leads in the target org

    Parameters:
        sf (Salesforce) -- the active Salesforce connection
        users (dict of string : string) -- the user to assign ownership of the Leads to
        wb (openpyxl.workbook.Workbook) -- the workbook containing the test data to create
        accounts (dict of string : string) -- the accounts to relate the Leads to

    Returns:
        void
    """
    ws = wb["Leads"]  # Gets the Leads sheet
    recordTypeMap = getRecordTypes(sf, ws, 1, "Lead")
    insertLeads = []
    try:
        logInfo("Reading Leads")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if (row[0] == None):
                continue
            insertLeads.append(
                {'RecordTypeId': recordTypeMap.get(row[0]), 'OwnerId': users.get(row[1]),
                 'Salutation': u"" if row[2] is None else row[2], 'FirstName': u"" if row[3] is None else row[3], 'LastName': u""
                 if row[4] is None else row[4], 'MiddleName': u"" if row[5] is None else row[5], 'Suffix': u"" if row[6] is None else row[6],
                 'EEP_Preferred_Name__c': u"" if row[7] is None else row[7], 'Company': row[8], 'EEP_Gender__c': row[9], 'Email': row[10],
                 'phone': row[11], 'MobilePhone': row[12], 'EEP_Preferred_Day__c': row[13],
                 'EEP_Producer_Account_Tax_Id__c': u"" if row[15] is None else row[15], 'EEP_National_Producer_Number__c': u"" if row[16] is None else row[16],
                 'EEP_Producer_CBU__c': row[17], 'EEP_Producer_Distribution_Channel__c': u"" if row[18] is None else row[18],
                 'Status': row[19], 'EEP_Closed_Lost_Reason__c': row[20],
                 'LeadSource': row[21], 'EEP_Source_Campaign__c': u"" if row[22] is None else row[22],
                 'EEP_Restricted_Access__c': row[23], 'EEP_Firm_Segment__c': row[24], 'HasOptedOutOfEmail': row[25],
                 'Street': u"" if row[26] is None else row[26], 'City': u"" if row[27] is None else row[27],
                 'State': u"" if row[28] is None else row[28], 'PostalCode': u"" if row[29] is None else row[29],
                 'Country': u"" if row[30] is None else row[30], 'FinServ__RelatedAccount__c': accounts.get(row[31]),
                 'FinServ__ReferredByUser__c': users.get(row[32]), 'EEP_Date_Of_Birth__c': "1970-05-09"})
            logging.info(insertLeads)
    except Exception as ex:
        logError("Could not read Leads", ex)

    try:
        logInfo("Creating Leads")
        Leads = sf.bulk.Lead.insert(insertLeads, batch_size=100)
        logInfo("Created Leads")
        logging.info(Leads)
    except Exception as ex:
        logError("Could not create Leads", ex)


def createOpportunities(sf, users, wb, accounts):
    """Creates Opportunities in the target org

    Parameters:
        sf (Salesforce) -- the active Salesforce connection
        users (dict of string : string) -- the user to assign ownership of the Opportunities to
        wb (openpyxl.workbook.Workbook) -- the workbook containing the test data to create
        accounts (dict of string : string) -- the accounts to relate the Opportunities to

    Returns:
        void
    """
    ws = wb["Opportunities"]  # Gets the Opportunities sheet
    recordTypeMap = getRecordTypes(sf, ws, 1, "Opportunity")
    insertOpportunities = []
    try:
        logInfo("Reading Opportunities")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if (row[0] == None):
                continue
            insertOpportunities.append(
                {'RecordTypeId': recordTypeMap.get(row[0]),
                 'OwnerId': users.get(row[1]),
                 'AccountId': accounts.get(row[2]),
                 'Name': row[3],
                 'Type': row[4],
                 'Budget_Confirmed__c': row[5],
                 'Discovery_Completed__c': row[6],
                 'ROI_Analysis_Completed__c': row[7],
                 'CloseDate': str(row[9]).replace(' 00:00:00', ''),
                 'StageName': row[10],
                 'Amount': 0 if row[12] is None else row[12],
                 'LeadSource': row[13],
                 'EEP_Producer_CBU__c': row[14],
                 'EEP_Producer_Distribution_Channel__c': row[15],
                 'EEP_Restricted_Access__c': row[16]})
            logging.info(insertOpportunities)
    except Exception as ex:
        logError("Could not read Opportunities", ex)

    try:
        logInfo("Creating Opportunities")
        Opportunities = sf.bulk.Opportunity.insert(insertOpportunities, batch_size=100)
        logInfo("Created Opportunities")
        logging.info(Opportunities)
    except Exception as ex:
        logError("Could not create Opportunities", ex)


def createTasks(sf, users, wb, accounts, contacts):
    """Creates Tasks in the target org

    Parameters:
        sf (Salesforce) -- the active Salesforce connection
        users (dict of string : string) -- the user to assign ownership of the Tasks to
        wb (openpyxl.workbook.Workbook) -- the workbook containing the test data to create
        accounts (dict of string : string) -- the accounts to relate the Tasks to
        contacts (dict of string : string) -- the contacts to relate the Tasks to

    Returns:
        void
    """
    ws = wb["Tasks"]
    insertTasks = []
    try:
        logInfo("Reading Tasks")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if (row[0] == None):
                continue
            insertTasks.append(
                {'Subject': row[0],
                 'Type': row[1],
                 'WhoId': contacts.get(row[2]),
                 # converts the date into a standardized datetime string then removes the time part due to the field only being a date field
                 'ActivityDate': str(row[3].isoformat()).replace('T00:00:00', ''),
                 'WhatId': accounts.get(row[4]),
                 'Priority': row[5],
                 'Status': row[6],
                 'OwnerId': users.get(row[9])})
            logging.info(insertTasks)
    except Exception as ex:
        logError("Could not read Tasks", ex)

    try:
        logInfo("Creating Tasks")
        Tasks = sf.bulk.Task.insert(insertTasks, batch_size=100)
        logInfo("Created Tasks")
        logging.info(Tasks)
    except Exception as ex:
        logError("Could not create Tasks", ex)


def createCases(sf, producers, wb, accounts, contacts):
    """Creates Cases in the target org

    Parameters:
        sf (Salesforce) -- the active Salesforce connection
        users (dict of string : string) -- the user to assign ownership of the Cases to
        wb (openpyxl.workbook.Workbook) -- the workbook containing the test data to create
        accounts (dict of string : string) -- the accounts to relate the Cases to
        contacts (dict of string : string) -- the contacts to relate the Cases to

    Returns:
        void
    """
    ws = wb["Cases"]
    insertCases = []
    try:
        logInfo("Reading Cases")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if (row[0] == None):
                continue
            insertCases.append(
                {'Type': row[0],
                 'Origin': row[1],
                 'EEP_Producer__c': producers.get(row[2]),
                 'ContactId': contacts.get(row[3]),
                 'Status': row[4],
                 'Priority': row[5],
                 'AccountId': accounts.get(row[6])})
            logging.info(insertCases)
    except Exception as ex:
        logError("Could not read Cases", ex)

    try:
        logInfo("Creating Cases")
        Cases = sf.bulk.Case.insert(insertCases, batch_size=100)
        logInfo("Created Cases")
        logging.info(Cases)
    except Exception as ex:
        logError("Could not create Cases", ex)


def createOperatingHours(sf):
    """Creates OperatingHours in the target org

    Parameters:
        sf (Salesforce) -- the active Salesforce connection

    Returns:
        OperatingHours (dict of string : string) -- a dictionary of the created OperatingHours with Name as the key and Id as the value
    """
    insertOperatingHours = []
    try:
        insertOperatingHours.append(
            {'Name': 'test hours',
            'timezone': 'America/New_York'
            })
        logging.info(insertOperatingHours)
    except Exception as ex:
        logError("Could not read OperatingHours", ex)
    try:
        logInfo("Creating OperatingHours")
        operatingHours = sf.bulk.OperatingHours.insert(insertOperatingHours, batch_size=100)
        logInfo("Created OperatingHours")
        logging.info(operatingHours)
        return queryCreatedRecords(sf, operatingHours, "OperatingHours")
    except Exception as ex:
        logError("Could not create OperatingHours", ex)


def createWorkType(sf, operatingHours):
    """Creates WorkType in the target org

    Parameters:
        sf (Salesforce) -- the active Salesforce connection
        operatingHours (dict of string : string) -- the operatingHours to relate the WorkType to

    Returns:
        WorkType (dict of string : string) -- a dictionary of the created WorkType with Name as the key and Id as the value
    """
    insertWorkType = []
    try:
        insertWorkType.append(
            {'Name': 'test work type',
            'OperatingHoursId': operatingHours.get('test hours'),
            'EstimatedDuration': 20,
            'DurationType': 'Hours'
            })
        logging.info(insertWorkType)
    except Exception as ex:
        logError("Could not read workType", ex)
    try:
        logInfo("Creating WorkType")
        workType = sf.bulk.WorkType.insert(insertWorkType, batch_size=100)
        logInfo("Created workType")
        logging.info(workType)
        return queryCreatedRecords(sf, workType, "WorkType")
    except Exception as ex:
        logError("Could not create WorkType", ex)


def createServiceTerritory(sf, operatingHours):
    """Creates ServiceTerritory in the target org

    Parameters:
        sf (Salesforce) -- the active Salesforce connection
        operatingHours (dict of string : string) -- the operatingHours to relate the ServiceTerritory to

    Returns:
        serviceTerritory (dict of string : string) -- a dictionary of the created serviceTerritory with Name as the key and Id as the value
    """
    insertServiceTerritory = []
    try:
        insertServiceTerritory.append(
            {'Name': 'test service territory',
            'OperatingHoursId': operatingHours.get('test hours'),
            'isActive': True,
            'Country': 'United States'
            })
        logging.info(insertServiceTerritory)
    except Exception as ex:
        logError("Could not read ServiceTerritory", ex)
    try:
        logInfo("Creating ServiceTerritory")
        serviceTerritory = sf.bulk.ServiceTerritory.insert(insertServiceTerritory, batch_size=100)
        logInfo("Created ServiceTerritory")
        logging.info(serviceTerritory)
        return queryCreatedRecords(sf, serviceTerritory, "serviceTerritory")
    except Exception as ex:
        logError("Could not create ServiceTerritory", ex)


def createServiceTerritoryWorkType(sf, serviceTerritory, workType):
    """Creates ServiceTerritory in the target org

    Parameters:
        sf (Salesforce) -- the active Salesforce connection
        serviceTerritory (dict of string : string) -- the serviceTerritory to relate the ServiceTerritoryWorkType to

    Returns:
        void
    """
    insertServiceTerritoryWorkType = []
    try:
        insertServiceTerritoryWorkType.append(
            {'ServiceTerritoryId': serviceTerritory.get('test service territory'),
            'WorkTypeId': workType.get('test work type'),
            })
        logging.info(insertServiceTerritoryWorkType)
    except Exception as ex:
        logError("Could not read ServiceTerritoryWorkType", ex)
    try:
        logInfo("Creating ServiceTerritoryWorkType")
        serviceTerritoryWorkType = sf.bulk.ServiceTerritoryWorkType.insert(insertServiceTerritoryWorkType, batch_size=100)
        logInfo("Created ServiceTerritoryWorkType")
        logging.info(serviceTerritoryWorkType)
    except Exception as ex:
        logError("Could not create ServiceTerritoryWorkType", ex)


def createWorkTypeGroup(sf):
    """Creates WorkTypeGroup in the target org

    Parameters:
        sf (Salesforce) -- the active Salesforce connection

    Returns:
        void
    """
    insertWorkTypeGroup = []
    try:
        insertWorkTypeGroup.append(
            {'Name': 'test work type group',
            'isActive': True,
            'GroupType': 'Default'
            })
        logging.info(insertWorkTypeGroup)
    except Exception as ex:
        logError("Could not read WorkTypeGroup", ex)
    try:
        logInfo("Creating WorkTypeGroup")
        WorkTypeGroup = sf.bulk.WorkTypeGroup.insert(insertWorkTypeGroup, batch_size=100)
        logInfo("Created WorkTypeGroup")
        logging.info(WorkTypeGroup)
    except Exception as ex:
        logError("Could not create WorkTypeGroup", ex)



if __name__ == '__main__':
    main()
